#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Teams & Customers Dashboard Builder
Reads operational reports and client data from Excel, generates v3_data.json,
then assembles dashboard_v7.html from components.

Usage:
    python3 build_data.py              # One-time build
    python3 build_data.py --watch      # Watch for changes, auto-rebuild
    python3 build_data.py --watch -i 15  # Custom interval (seconds)
"""

import os
import sys
import json
import time
import shutil
import tempfile
import argparse
from datetime import datetime
from collections import defaultdict

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# === PATHS ===
SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ONEDRIVE_BASE = os.path.expanduser("~/Library/CloudStorage/OneDrive-AscorpSP/My Obsidian/FinancesDocs")
OPS_EXCEL = os.path.join(ONEDRIVE_BASE, "Операционные отчеты (ежедневные).xlsx")
CLIENT_EXCEL = os.path.join(ONEDRIVE_BASE, "Отчет по клиентам (ежемесячный).xlsx")
JSON_OUTPUT = os.path.join(SCRIPT_DIR, "v3_data.json")
HTML_OUTPUT = os.path.join(SCRIPT_DIR, "dashboard_v7.html")
CSS_PATH = os.path.join(SCRIPT_DIR, "new_css.txt")
BODY_PATH = os.path.join(SCRIPT_DIR, "new_body.txt")
JS_PATH = os.path.join(SCRIPT_DIR, "new_js.txt")
LEADERS_DIR = os.path.expanduser("~/Library/CloudStorage/OneDrive-AscorpSP/Leaders Dashboards")
LEADERS_HTML = os.path.join(LEADERS_DIR, "teams-customers-dashboard.html")

# === CONFIGURATION ===

# Sheet name → subgroup configuration
# гамма-1 sheet contains 4 subgroups separated by "Итого" rows
# All other sheets map to a single subgroup
SHEET_CONFIG = {
    'гамма-1': {
        'subgroups': ['Гамма-1', 'Гамма-1А', 'Гамма-1Б', 'Гамма-1 Полевые'],
        'group': 'Гамма',
    },
    'альфа-1': {'subgroups': ['Альфа-1'], 'group': 'Альфа'},
    'альфа-2': {'subgroups': ['Альфа-2'], 'group': 'Альфа'},
    'гамма-2': {'subgroups': ['Гамма-2'], 'group': 'Гамма'},
    'дельта': {'subgroups': ['Дельта'], 'group': 'Дельта'},
    'вита': {'subgroups': ['Вита'], 'group': 'Вита'},
    'тета': {'subgroups': ['Тета'], 'group': 'Тета'},
    'дзета': {'subgroups': ['Дзета'], 'group': 'Дзета'},
}

# Russian month names
MONTH_NAMES = {
    1: 'январь', 2: 'февраль', 3: 'март', 4: 'апрель',
    5: 'май', 6: 'июнь', 7: 'июль', 8: 'август',
    9: 'сентябрь', 10: 'октябрь', 11: 'ноябрь', 12: 'декабрь',
}
MONTH_NAME_TO_NUM = {v: k for k, v in MONTH_NAMES.items()}

# Russian production calendar: working days per month
PROD_CALENDAR = {
    'январь 2025': 17, 'февраль 2025': 19, 'март 2025': 20,
    'апрель 2025': 22, 'май 2025': 17, 'июнь 2025': 19,
    'июль 2025': 23, 'август 2025': 21, 'сентябрь 2025': 22,
    'октябрь 2025': 23, 'ноябрь 2025': 19, 'декабрь 2025': 22,
    'январь 2026': 15, 'февраль 2026': 19, 'март 2026': 22,
    'апрель 2026': 22, 'май 2026': 18, 'июнь 2026': 21,
    'июль 2026': 23, 'август 2026': 21, 'сентябрь 2026': 22,
    'октябрь 2026': 22, 'ноябрь 2026': 20, 'декабрь 2026': 22,
}

# Column header → field name mapping (normalized: stripped, lowercase)
HEADER_MAP = {
    'решенные заявки': 'tk_b',
    'решенные задачи': 'ts_b',
    'решенные рег.заявки': 'tk_r',
    'решенные рег.задачи': 'ts_r',
    'рег. заявки /задачи': '_reg_combined',
    'рег.заявки/задачи': '_reg_combined',
    'решенные заявки и задачи': '_reg_combined',
    'решенные заявки/задачи': '_reg_combined',
    'решенные заявки/\nзадачи': '_reg_combined',
    'выезды': 'vz',
    'решенные зни': '_zni',
    'тзт': 'tzt',
}

# Row labels to skip in operational data
SKIP_NAMES = {
    'итого', 'беклог', 'сотрудники', '',
    'беклог (0-2)', 'бектог (2-4)', 'беклог (2-4)',
    'беклог (5-10)', 'беклог (больше 4)', 'беклог (больше 10)',
    'беклог (без проектов)',
}


# === HELPERS ===

def safe_load_workbook(path, max_retries=3, retry_delay=5):
    """Load Excel workbook via a temporary copy for safety.

    Copies the file to a temp location before reading to avoid any interaction
    with the original file (OneDrive sync, Excel locks, etc.).
    Retries on failure (e.g. file being synced).
    """
    for attempt in range(1, max_retries + 1):
        tmp_path = None
        try:
            tmp_fd, tmp_path = tempfile.mkstemp(suffix='.xlsx')
            os.close(tmp_fd)
            shutil.copy2(path, tmp_path)
            wb = openpyxl.load_workbook(tmp_path, data_only=True)
            return wb
        except Exception as e:
            if attempt < max_retries:
                print(f"  Retry {attempt}/{max_retries} loading {os.path.basename(path)}: {e}")
                time.sleep(retry_delay)
            else:
                raise
        finally:
            if tmp_path and os.path.exists(tmp_path):
                try:
                    os.unlink(tmp_path)
                except OSError:
                    pass


def safe_float(v):
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    try:
        return float(str(v).strip())
    except (ValueError, TypeError):
        return 0.0


def safe_int(v):
    return int(round(safe_float(v)))


def parse_date(s):
    if not s:
        return None
    try:
        return datetime.strptime(str(s).strip(), "%d.%m.%Y")
    except ValueError:
        return None


def month_label(dt):
    return f"{MONTH_NAMES[dt.month]} {dt.year}"


def month_sort_key(ml):
    parts = ml.split()
    return (int(parts[1]), MONTH_NAME_TO_NUM[parts[0]])


def find_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    for sn in wb.sheetnames:
        if sn.strip().lower() == name.strip().lower():
            return wb[sn]
    return None


# === OPERATIONAL DATA PARSING ===

def detect_block_columns(ws, header_row):
    """Detect column layout from a block's header row.

    Returns list of field names, one per column offset from the date start.
    The stride = len(field_list).
    """
    fields = []
    for c in range(2, min(ws.max_column + 1, 200)):
        v = ws.cell(header_row, c).value
        if v is None:
            break
        h = str(v).strip().lower()
        mapped = HEADER_MAP.get(h)
        if mapped is None:
            break
        # Detect repeat = end of one date group
        if fields and mapped == fields[0]:
            break
        fields.append(mapped)
    if not fields:
        fields = ['tk_b', 'ts_b', 'tk_r', 'ts_r', 'tzt']

    # Post-process: resolve _reg_combined and _zni based on context
    has_separate_tsr = 'ts_r' in fields
    result = []
    for i, f in enumerate(fields):
        if f == '_reg_combined':
            # Combined "рег. заявки/задачи" column → always map to tk_r.
            # This is a sum of tk_r + ts_r that can't be split;
            # ts_r should only come from a dedicated "решенные рег.задачи" column.
            result.append('tk_r')
        elif f == '_zni':
            # зни → ts_r only when there's no separate ts_r column
            result.append('ts_r' if not has_separate_tsr else '_zni')
        else:
            result.append(f)
    return result


def parse_ops_sheet(wb, sheet_name, config):
    ws = find_sheet(wb, sheet_name)
    if not ws:
        print(f"  WARNING: Sheet '{sheet_name}' not found")
        return []

    subgroups = config['subgroups']
    group = config['group']
    daily_records = []

    # Find all block boundaries ("Сотрудники" in col A)
    block_starts = []
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if v and str(v).strip() == 'Сотрудники':
            block_starts.append(r)

    for bi, bs in enumerate(block_starts):
        date_row = bs + 2
        data_start = bs + 4
        block_end = block_starts[bi + 1] - 1 if bi + 1 < len(block_starts) else ws.max_row

        # Validate: first date cell must parse
        if not parse_date(ws.cell(date_row, 2).value):
            continue

        # Detect column layout for this block
        fields = detect_block_columns(ws, bs)
        stride = len(fields)

        # Collect dates
        dates = []
        for c in range(2, ws.max_column + 1, stride):
            dt = parse_date(ws.cell(date_row, c).value)
            if dt:
                dates.append((c, dt))
        if not dates:
            continue

        # Parse employee rows; track subgroup via gaps (blank rows, Итого, Беклог)
        # In гамма-1, subgroups are separated by blank/Итого/Беклог rows.
        # First subgroup (Гамма-1) has no trailing Итого — just a blank row before next.
        sg_idx = 0
        in_gap = False
        for r in range(data_start, block_end + 1):
            name_val = ws.cell(r, 1).value
            if not name_val:
                in_gap = True
                continue
            name = str(name_val).strip()
            name_lower = name.lower()

            if name_lower in SKIP_NAMES:
                in_gap = True
                continue

            # Employee row: advance subgroup if we were in a gap
            if in_gap and sg_idx < len(subgroups) - 1:
                sg_idx += 1
            in_gap = False

            sg = subgroups[min(sg_idx, len(subgroups) - 1)]

            for col_start, dt in dates:
                rec = {
                    'e': name, 'sg': sg, 'g': group,
                    'd': dt.strftime('%Y-%m-%d'),
                    'm': month_label(dt),
                    'tzt': 0.0, 'tk_b': 0, 'ts_b': 0,
                    'tk_r': 0, 'ts_r': 0, 'vz': 0,
                }
                for offset, field in enumerate(fields):
                    val = ws.cell(r, col_start + offset).value
                    if field == 'tzt':
                        rec['tzt'] = round(safe_float(val), 2)
                    elif field == 'vz':
                        v = safe_int(val)
                        rec['vz'] = v
                        if 'ts_r' not in fields:
                            rec['ts_r'] = v  # count as ts_r only when no dedicated ts_r column
                    elif field in ('tk_b', 'ts_b', 'tk_r', 'ts_r'):
                        rec[field] = safe_int(val)
                    # _zni (when separate ts_r exists) and unknowns are skipped
                daily_records.append(rec)

    return daily_records


def parse_all_ops(ops_path):
    print(f"Parsing: {os.path.basename(ops_path)}")
    wb = safe_load_workbook(ops_path)
    all_daily = []
    for sheet_name, config in SHEET_CONFIG.items():
        records = parse_ops_sheet(wb, sheet_name, config)
        print(f"  {sheet_name}: {len(records)} daily records")
        all_daily.extend(records)
    print(f"  Total: {len(all_daily)} daily records")
    return all_daily


# === AGGREGATION ===

def build_hierarchy(daily):
    sg_employees = defaultdict(set)
    sg_to_group = {}
    for r in daily:
        sg_employees[r['sg']].add(r['e'])
        sg_to_group[r['sg']] = r['g']

    hierarchy = {}
    group_map = {}
    # Preserve SHEET_CONFIG order for groups and subgroups
    seen_groups = []
    for cfg in SHEET_CONFIG.values():
        g = cfg['group']
        if g not in seen_groups:
            seen_groups.append(g)
    for g in seen_groups:
        hierarchy[g] = {}
        group_map[g] = []
    for cfg in SHEET_CONFIG.values():
        g = cfg['group']
        for sg in cfg['subgroups']:
            if sg in sg_employees:
                hierarchy[g][sg] = sorted(sg_employees[sg])
                if sg not in group_map[g]:
                    group_map[g].append(sg)

    return hierarchy, group_map, sg_to_group


def compute_months(daily):
    months = sorted(set(r['m'] for r in daily), key=month_sort_key)
    mo_map = {m: i + 1 for i, m in enumerate(months)}
    return months, mo_map


def aggregate(daily, ops_mo_map):
    emp_agg = defaultdict(lambda: {
        'tzt': 0.0, 'tk_b': 0, 'ts_b': 0, 'tk_r': 0, 'ts_r': 0, 'vz': 0,
    })
    emp_meta = {}
    sg_agg = defaultdict(lambda: {
        'tzt': 0.0, 'tk_b': 0, 'ts_b': 0, 'tk_r': 0, 'ts_r': 0, 'vz': 0,
        'employees': set(), 'dates': set(),
    })
    sg_meta = {}

    for r in daily:
        ek = (r['e'], r['sg'], r['m'])
        d = emp_agg[ek]
        d['tzt'] += r['tzt']
        d['tk_b'] += r['tk_b']
        d['ts_b'] += r['ts_b']
        d['tk_r'] += r['tk_r']
        d['ts_r'] += r['ts_r']
        d['vz'] += r['vz']
        emp_meta[ek] = r['g']

        sk = (r['sg'], r['m'])
        s = sg_agg[sk]
        s['tzt'] += r['tzt']
        s['tk_b'] += r['tk_b']
        s['ts_b'] += r['ts_b']
        s['tk_r'] += r['tk_r']
        s['ts_r'] += r['ts_r']
        s['vz'] += r['vz']
        s['employees'].add(r['e'])
        s['dates'].add(r['d'])
        sg_meta[sk] = r['g']

    # Employee monthly
    emp_monthly = []
    for (emp, sg, month), d in sorted(emp_agg.items()):
        tzt = round(d['tzt'], 1)
        norm = PROD_CALENDAR.get(month, 21) * 8
        util = round(tzt / norm * 100, 1) if norm > 0 else 0
        tph_b = round(d['tk_b'] / tzt, 4) if tzt > 0 else 0
        tph_z = round(d['ts_b'] / tzt, 4) if tzt > 0 else 0
        tph_all = round((d['tk_b'] + d['ts_b']) / tzt, 4) if tzt > 0 else 0
        emp_monthly.append({
            'employee': emp, 'subgroup': sg, 'group': emp_meta[(emp, sg, month)],
            'month': month, 'month_order': ops_mo_map[month],
            'tzt': tzt, 'norm': norm, 'util': util,
            'tk_b': d['tk_b'], 'ts_b': d['ts_b'],
            'tk_r': d['tk_r'], 'ts_r': d['ts_r'],
            'tph_b': tph_b, 'tph_z': tph_z, 'tph_all': tph_all,
            'vz': d['vz'],
        })

    # Subgroup monthly
    sg_monthly = []
    for (sg, month), s in sorted(sg_agg.items()):
        num_emp = len(s['employees'])
        tzt = round(s['tzt'], 1)
        norm = PROD_CALENDAR.get(month, 21) * 8 * num_emp
        util = round(tzt / norm * 100, 1) if norm > 0 else 0
        tph_b = round(s['tk_b'] / tzt, 4) if tzt > 0 else 0
        tph_z = round(s['ts_b'] / tzt, 4) if tzt > 0 else 0
        tph_all = round((s['tk_b'] + s['ts_b']) / tzt, 4) if tzt > 0 else 0
        sg_monthly.append({
            'subgroup': sg, 'group': sg_meta[(sg, month)],
            'month': month, 'month_order': ops_mo_map[month],
            'tzt': tzt, 'norm': norm, 'util': util,
            'tk_b': s['tk_b'], 'ts_b': s['ts_b'],
            'tk_r': s['tk_r'], 'ts_r': s['ts_r'],
            'employees': num_emp, 'days': len(s['dates']),
            'tph_b': tph_b, 'tph_z': tph_z, 'tph_all': tph_all,
            'vz': s['vz'],
        })

    return emp_monthly, sg_monthly


# === CLIENT DATA PARSING ===

def infer_month_years(month_names):
    """Assign year to each month name based on chronological order.
    Assumes months run forward; year increments when month number decreases.
    """
    years = {}
    year = 2025
    prev_num = 0
    for mn in month_names:
        num = MONTH_NAME_TO_NUM.get(mn, 0)
        if num and num < prev_num:
            year += 1
        prev_num = num
        years[mn] = year
    return years


def parse_cl_tzt(wb):
    ws = wb['данные тзт']
    results = []
    canonical_clients = {}  # lowercase → canonical name

    for r in range(2, ws.max_row + 1):
        client = ws.cell(r, 1).value
        month_name = ws.cell(r, 2).value
        year = ws.cell(r, 3).value
        team = ws.cell(r, 4).value
        tzt_type = ws.cell(r, 5).value
        tzt = ws.cell(r, 6).value

        if not client or not month_name:
            continue

        client = str(client).strip()
        canonical_clients[client.lower()] = client
        month_name = str(month_name).strip().lower()
        year = int(year) if year else 2025
        team = str(team).strip() if team else ''
        tzt_type = str(tzt_type).strip() if tzt_type else 'операционка'

        ml = f"{month_name} {year}"
        results.append({
            'client': client, 'month': month_name, 'ml': ml,
            'team': team, 'tzt_type': tzt_type,
            'tzt': round(safe_float(tzt), 2),
        })

    return results, canonical_clients


# Known aliases for client names that differ between Excel sheets
CLIENT_ALIASES = {
    'самитагро': 'Самми Агро',
    'карабанов': 'Карабанов и партнеры',
    'сева': 'Ceva',
    'судьи': 'Мировые судьи',
    'кистоун': 'Кистоун Лоджистикс',
}


def normalize_client(name, canonical_map):
    """Normalize client name to canonical form."""
    key = name.lower()
    if key in canonical_map:
        return canonical_map[key]
    if key in CLIENT_ALIASES:
        return CLIENT_ALIASES[key]
    return name


def parse_cl_tickets(wb, sheet_name, ticket_type, canonical_clients):
    """Parse заявки or задачи sheet (pivot format: months × поступило/решено)."""
    ws = wb[sheet_name]
    results = []

    # Row 1: month names at even columns
    months = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and str(v).strip().lower() in MONTH_NAME_TO_NUM:
            months.append((c, str(v).strip().lower()))

    if not months:
        return results

    month_years = infer_month_years([m for _, m in months])
    # Stride = distance between month columns
    stride = months[1][0] - months[0][0] if len(months) >= 2 else 2

    # Data starts at row 3
    for r in range(3, ws.max_row + 1):
        client = ws.cell(r, 1).value
        if not client or not str(client).strip():
            continue
        client = normalize_client(str(client).strip(), canonical_clients)

        for col_start, mn in months:
            year = month_years.get(mn, 2025)
            ml = f"{mn} {year}"
            incoming = safe_int(ws.cell(r, col_start).value)
            resolved = safe_int(ws.cell(r, col_start + 1).value)
            results.append({
                'client': client, 'month': mn, 'ml': ml,
                'type': ticket_type,
                'incoming': incoming, 'resolved': resolved,
            })

    return results


def parse_cl_sla(wb, canonical_clients):
    ws = wb['sla']
    results = []

    months = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and str(v).strip().lower() in MONTH_NAME_TO_NUM:
            months.append((c, str(v).strip().lower()))

    if not months:
        return results

    month_years = infer_month_years([m for _, m in months])
    stride = months[1][0] - months[0][0] if len(months) >= 2 else 2

    for r in range(3, ws.max_row + 1):
        client = ws.cell(r, 1).value
        if not client or not str(client).strip():
            continue
        client = normalize_client(str(client).strip(), canonical_clients)

        for col_start, mn in months:
            year = month_years.get(mn, 2025)
            ml = f"{mn} {year}"
            sr_raw = ws.cell(r, col_start).value
            si_raw = ws.cell(r, col_start + 1).value
            sr = safe_float(sr_raw) if sr_raw and str(sr_raw).strip() != '-' else None
            si = safe_float(si_raw) if si_raw and str(si_raw).strip() != '-' else None
            results.append({
                'client': client, 'month': mn, 'ml': ml,
                'sr': sr, 'si': si,
            })

    return results


def parse_cl_mass(wb, canonical_clients):
    ws = wb['массовые']
    results = []

    # Row 1: month names, 1 column per month
    months = []
    for c in range(2, ws.max_column + 1):
        v = ws.cell(1, c).value
        if v and str(v).strip().lower() in MONTH_NAME_TO_NUM:
            months.append((c, str(v).strip().lower()))

    if not months:
        return results

    month_years = infer_month_years([m for _, m in months])

    # Data starts at row 2 (no sub-header row)
    for r in range(2, ws.max_row + 1):
        client = ws.cell(r, 1).value
        if not client or not str(client).strip():
            continue
        client = normalize_client(str(client).strip(), canonical_clients)

        for col, mn in months:
            year = month_years.get(mn, 2025)
            ml = f"{mn} {year}"
            mi = safe_int(ws.cell(r, col).value)
            results.append({
                'client': client, 'month': mn, 'ml': ml, 'mi': mi,
            })

    return results


def parse_all_client(client_path):
    print(f"Parsing: {os.path.basename(client_path)}")
    wb = safe_load_workbook(client_path)

    cl_tzt, canonical_clients = parse_cl_tzt(wb)
    print(f"  данные тзт: {len(cl_tzt)} records")

    cl_tickets = parse_cl_tickets(wb, 'заявки', 'заявки', canonical_clients)
    cl_tasks = parse_cl_tickets(wb, 'задачи', 'задачи', canonical_clients)
    cl_all_tickets = cl_tickets + cl_tasks
    print(f"  заявки: {len(cl_tickets)}, задачи: {len(cl_tasks)}")

    cl_sla = parse_cl_sla(wb, canonical_clients)
    print(f"  sla: {len(cl_sla)} records")

    cl_mass = parse_cl_mass(wb, canonical_clients)
    print(f"  массовые: {len(cl_mass)} records")

    return cl_tzt, cl_all_tickets, cl_sla, cl_mass


# === BUILD ===

def build_data():
    # 1. Operational data
    daily = parse_all_ops(OPS_EXCEL)
    hierarchy, group_map, sg_to_group = build_hierarchy(daily)
    months_ops, ops_mo_map = compute_months(daily)

    for r in daily:
        r['mo'] = ops_mo_map[r['m']]

    emp_monthly, sg_monthly = aggregate(daily, ops_mo_map)

    # 2. Client data
    cl_tzt, cl_tickets, cl_sla, cl_mass = parse_all_client(CLIENT_EXCEL)

    # Client month ordering
    cl_months = sorted(set(r['ml'] for r in cl_tzt), key=month_sort_key)
    cl_mo_map = {m: i + 1 for i, m in enumerate(cl_months)}

    for dataset in (cl_tzt, cl_tickets, cl_sla, cl_mass):
        for r in dataset:
            r['mo'] = cl_mo_map.get(r['ml'], 0)

    # Clients: collect from ALL client data sources (some clients appear only in tickets/sla)
    all_client_names = set(r['client'] for r in cl_tzt)
    all_client_names |= set(r['client'] for r in cl_tickets)
    all_client_names |= set(r['client'] for r in cl_sla)
    all_client_names |= set(r['client'] for r in cl_mass)
    clients = sorted(all_client_names)
    teams_cl = sorted(set(r['team'] for r in cl_tzt))

    # 3. Production calendar for all relevant months
    all_months = sorted(set(months_ops) | set(cl_months), key=month_sort_key)
    prod_cal = {m: PROD_CALENDAR.get(m, 21) for m in all_months}

    # 4. Assemble
    data = {
        'hierarchy': hierarchy,
        'group_map': group_map,
        'sg_to_group': sg_to_group,
        'months_ops': months_ops,
        'months_cl': cl_months,
        'clients': clients,
        'teams_cl': teams_cl,
        'daily': daily,
        'sg_monthly': sg_monthly,
        'emp_monthly': emp_monthly,
        'cl_tzt': cl_tzt,
        'cl_tickets': cl_tickets,
        'cl_sla': cl_sla,
        'cl_mass': cl_mass,
        'prod_calendar': prod_cal,
    }

    print(f"\nData summary:")
    print(f"  {len(daily)} daily, {len(emp_monthly)} emp_monthly, {len(sg_monthly)} sg_monthly")
    print(f"  {len(cl_tzt)} cl_tzt, {len(cl_tickets)} cl_tickets, {len(cl_sla)} cl_sla, {len(cl_mass)} cl_mass")
    print(f"  Ops months: {months_ops}")
    print(f"  Client months: {cl_months}")
    print(f"  Teams: {len(hierarchy)} groups, {len(sg_to_group)} subgroups")
    print(f"  Clients: {len(clients)}, Client teams: {len(teams_cl)}")

    return data


def write_json(data):
    with open(JSON_OUTPUT, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, separators=(',', ':'))
    size_kb = os.path.getsize(JSON_OUTPUT) / 1024
    print(f"JSON: {JSON_OUTPUT} ({size_kb:.0f} KB)")


def build_html():
    with open(CSS_PATH, 'r', encoding='utf-8') as f:
        css = f.read()
    with open(BODY_PATH, 'r', encoding='utf-8') as f:
        body = f.read()
    with open(JS_PATH, 'r', encoding='utf-8') as f:
        js = f.read()
    with open(JSON_OUTPUT, 'r', encoding='utf-8') as f:
        data_json = f.read()

    html = f"""<!DOCTYPE html>
<html lang="ru">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width, initial-scale=1.0">
<meta http-equiv="refresh" content="60">
<title>Информационная панель</title>
<script src="https://cdn.jsdelivr.net/npm/chart.js"></script>
<style>
{css}
</style>
</head>
{body}
<script>
const D = {data_json};
{js}
</script>
</html>"""

    with open(HTML_OUTPUT, 'w', encoding='utf-8') as f:
        f.write(html)
    size_kb = os.path.getsize(HTML_OUTPUT) / 1024
    print(f"HTML: {HTML_OUTPUT} ({size_kb:.0f} KB)")


def copy_to_leaders():
    """Copy dashboard HTML to OneDrive Leaders Dashboards for sharing."""
    try:
        os.makedirs(LEADERS_DIR, exist_ok=True)
        shutil.copy2(HTML_OUTPUT, LEADERS_HTML)
        print(f"Leaders: copied to {LEADERS_HTML}")
    except Exception as e:
        print(f"Leaders: copy failed — {e}")


def build():
    start = time.time()
    data = build_data()
    write_json(data)
    build_html()
    copy_to_leaders()
    elapsed = time.time() - start
    print(f"\nBuild complete in {elapsed:.1f}s")


def watch(interval=30):
    files = [OPS_EXCEL, CLIENT_EXCEL]
    print(f"Watching {len(files)} files (interval: {interval}s)")
    print("Press Ctrl+C to stop\n")

    build()

    last_mtimes = {}
    for f in files:
        try:
            last_mtimes[f] = os.path.getmtime(f)
        except OSError:
            last_mtimes[f] = 0

    while True:
        try:
            time.sleep(interval)
            changed = False
            for f in files:
                try:
                    mt = os.path.getmtime(f)
                    if mt != last_mtimes.get(f, 0):
                        changed = True
                        last_mtimes[f] = mt
                except OSError:
                    continue

            if changed:
                print(f"\n{'=' * 60}")
                print(f"Change detected at {datetime.now().strftime('%H:%M:%S')}")
                print(f"{'=' * 60}")
                try:
                    build()
                except Exception as e:
                    print(f"Build error: {e}")

        except KeyboardInterrupt:
            print("\nStopped.")
            break


if __name__ == '__main__':
    parser = argparse.ArgumentParser(description='Teams & Customers Dashboard Builder')
    parser.add_argument('--watch', '-w', action='store_true', help='Watch for file changes')
    parser.add_argument('--interval', '-i', type=int, default=30, help='Watch interval (seconds)')
    args = parser.parse_args()

    for f in [OPS_EXCEL, CLIENT_EXCEL]:
        if not os.path.exists(f):
            print(f"ERROR: File not found: {f}")
            sys.exit(1)

    if args.watch:
        watch(args.interval)
    else:
        build()
